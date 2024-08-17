import os
import torch
from datasets import load_dataset
from torch.utils.data import DataLoader
from transformers import AutoModelForImageClassification
from transformers import CvtForImageClassification, DeiTForImageClassification, ViTForImageClassification, ViTHybridForImageClassification, ConvNextV2Model
# Model paths and their respective image processor imports
from transformers import DeiTImageProcessor, ViTImageProcessor, AutoImageProcessor, ViTHybridImageProcessor
import PIL
from PIL import Image
from datetime import datetime
import numpy

#####################################################################################################
#### diese Parameter müssen angepasst werden, um das Skript selbst auszuführen ####


#pls select the right split
ds = "/home/student_01/Projects/MA_Cahide/Datasets/Base/20_80_10_10_min4/test"
print(ds)

#Select the Models that you want to use as base for evaluation

model_info = {
    'ViTHybrid_30_SD': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/30/ViTHybrid_30_SD',
        'processor': ViTHybridImageProcessor.from_pretrained("google/vit-hybrid-base-bit-384"),
        'custom_class': ViTHybridForImageClassification
    },

    'ViTHybrid_30_GAN': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/30/ViTHybrid_30_GAN',
        'processor': ViTHybridImageProcessor.from_pretrained("google/vit-hybrid-base-bit-384"),
        'custom_class': ViTHybridForImageClassification
    },
    
    
    'ViTHybrid_30_all': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/30/ViTHybrid_30_all',
        'processor': ViTHybridImageProcessor.from_pretrained("google/vit-hybrid-base-bit-384"),
        'custom_class': ViTHybridForImageClassification
    }
    


}


#####################################################################################################


device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

print(f"{datetime.now()}:Using {device} for computation.")

# Load dataset

print(f"{datetime.now()}:~~~loading dataset~~~")
base_dir = ds
test_dataset = load_dataset('imagefolder', data_dir=base_dir, split="train")
print(test_dataset)

#print(test_dataset["image"])
labels = test_dataset.features["label"].names
id2label={str(i): c for i, c in enumerate(labels)}
label2id={c: str(i) for i, c in enumerate(labels)}


# Load Models
print(f"{datetime.now()}:~~~loading models~~~")
# Model information with their paths and processors



# Initialize models and their specific processors
models = {}
for name, info in model_info.items():
    print(f"{datetime.now()}:~~~loading { name }~~~")
    if 'custom_class' in info:  # Check if a custom class is specified
        model_class = info['custom_class']
    else:
        model_class = AutoModelForImageClassification
    models[name] = {
        'model': model_class.from_pretrained(info['path']).to(device),
        'processor': info['processor']
    }

output_dir="output/"
os.makedirs(output_dir, exist_ok=True)
##################################################################################################################################################################



image_paths = []

# Walk through the directory structure
for root, dirs, files in os.walk(base_dir):
    for file in files:
        # Check if the file is a JPEG image
        if file.lower().endswith('.jpg') or file.lower().endswith('.jpeg'):
            # Construct the relative path from the base directory
            relative_path = os.path.relpath(root, base_dir)
            # Construct the full path to the image file
            image_path = os.path.join(relative_path, file)
            # Append the path in the format 'classname/imagefile'
            image_paths.append(image_path)

# Now, image_paths contains all the image paths in the desired format
#print(image_paths)
#print(len(image_paths))


# Function to load and process an image

def process_image(image_input, processor):
    # Check if the input is a file path (str) and needs to be opened
    if isinstance(image_input, str):
        image = Image.open(image_input).convert("RGB")
    # Or if it's already an Image object
    elif isinstance(image_input, Image.Image):
        image = image_input.convert("RGB")
    else:
        raise TypeError("The image input must be a file path (str) or a PIL.Image object")

    # Process the image with the provided processor
    return processor(images=image, return_tensors="pt")


# Function to preprocess and get predictions from a model
def get_predictions_for_model(model, dataset, processor):
    model.eval()
    predictions = []
    with torch.no_grad():
        for img_path in dataset['image']:
            processed = process_image(img_path, processor)
            pixel_values = processed['pixel_values'].to(device)
            outputs = model(pixel_values=pixel_values)
            predictions.append(outputs.logits)
    return torch.cat(predictions, dim=0)

# Ensemble prediction logic
# Step 1: Store individual model predictions with identifiers
model_predictions_dict = {}
for name, info in models.items():
    print(f"{datetime.now()}: Processing {name}")
    predictions = get_predictions_for_model(info['model'], test_dataset, info['processor'])
    model_predictions_dict[name] = predictions

print(model_predictions_dict)

# Step 2: Calculate ensemble average
ensemble_predictions = torch.stack(list(model_predictions_dict.values())).mean(dim=0)
###
from scipy.stats import mode


# Step 1: Convert logits to class predictions for each model
class_predictions = {name: torch.argmax(preds, dim=1) for name, preds in model_predictions_dict.items()}

# Step 2: Stack the predictions into a single tensor for easier processing
# Assuming the number of inputs is N, this creates a tensor of shape [num_models, N]
stacked_predictions = torch.stack(list(class_predictions.values()))

# Step 3: Determine the most frequent class prediction (mode) for each input
# We use the scipy mode function because PyTorch does not have a built-in mode function.
# Note: mode().mode returns a numpy array, so we convert it back to a tensor.
# If you encounter dimensionality issues, ensure that the array is squeezed or reshaped appropriately before conversion.
mode_predictions = mode(stacked_predictions.cpu().numpy(), axis=0)[0]  # axis=0 to calculate mode across models
mode_predictions = torch.tensor(mode_predictions.squeeze(), device=device)  # Ensure tensor is on the correct device


# Continue with the accuracy calculation as before
true_labels = torch.tensor(test_dataset['label']).to(device)  # Ensure true_labels is on the same device as mode_predictions

# Calculate accuracy
accuracy = (mode_predictions == true_labels).float().mean()
print(f"{datetime.now()}: Accuracy with Hard Voting: {accuracy.item() * 100:.2f}%")
####


##################################################################################################################################################################

print(f"{datetime.now()}:~~~average predictions~~~")    
# Average the predictions from all models
#ensemble_predictions = torch.stack(ensemble_predictions).mean(dim=0)

# Assuming a classification task, get the final decision for each sample
final_decisions = torch.argmax(ensemble_predictions, dim=1)

true_labels = torch.tensor(test_dataset['label']).to(device)  #  to the same device as final_decisions

#  calculate accuracy
accuracy = (final_decisions == true_labels).float().mean()
print(f"{datetime.now()}: Accuracy: {accuracy.item() * 100:.2f}%")

from sklearn.metrics import f1_score

# Convert tensors to CPU for sklearn compatibility
true_labels_cpu = true_labels.cpu().numpy()
final_decisions_cpu = final_decisions.cpu().numpy()

# Calculate F1 score
f1 = f1_score(true_labels_cpu, final_decisions_cpu, average='weighted')  # Use 'average' parameter as per your requirement
print(f"{datetime.now()}: F1 Score: {f1 * 100:.2f}%")


##################################################################################################################################################################

print(f"{datetime.now()}:~~~visualise~~~")

from sklearn.metrics import confusion_matrix
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np
#test_dataset.features["label"].names
# Assuming true_labels and final_decisions are already defined and converted to CPU numpy arrays
cm = confusion_matrix(true_labels_cpu, final_decisions_cpu)


plt.figure(figsize=(30, 30))
mask = np.zeros_like(cm, dtype=bool)
mask[cm == 0] = True
sns.set(font_scale=1.0)
# Uncomment the following line to plot the absolute confusion matrix
# ax = sns.heatmap(cm, annot=mask, fmt="d", cmap="RdYlGn", xticklabels=labels, yticklabels=labels)
ax = sns.heatmap(cm, annot=True, fmt="d", cmap="RdYlGn", mask=mask, cbar=False, xticklabels=labels, yticklabels=labels)
    
ax.set_xlabel('Predicted labels', fontsize=15)
ax.set_ylabel('True labels', fontsize=15)
ax.set_title(f'Confusion Matrix for Hard Voter', fontsize=20)
ax.tick_params(axis='x', labelrotation=45, labelsize=10)
ax.tick_params(axis='y', labelrotation=0, labelsize=10)
    
# Save confusion matrix
plt.tight_layout()
plt.savefig("output/hard_voter_30_confusion_matrix.png", dpi=300)
plt.close()
print(f"{datetime.now()}: Confusion matrix for Hard Voter saved.")

# Calculate normalized confusion matrix
cm_normalized = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]

# Create a mask to hide the zeroes in the normalized confusion matrix
mask = cm_normalized == 0
    
# Plot normalized confusion matrix
plt.figure(figsize=(30, 30))  # Adjusted figure size for better readability
sns.set(font_scale=1.0)  # Smaller font scale for better readability
ax = sns.heatmap(cm_normalized, annot=~mask, fmt=".0f", cmap="RdYlGn", mask=mask, cbar=True, 
                     xticklabels=labels, yticklabels=labels)
ax.set_xlabel('Predicted labels', fontsize=15)
ax.set_ylabel('True labels', fontsize=15)
ax.set_title(f'Confusion Matrix for Hard Voter', fontsize=18)
ax.tick_params(axis='x', labelrotation=90, labelsize=10)
ax.tick_params(axis='y', labelrotation=0, labelsize=10)
    
# Save normalized confusion matrix
plt.tight_layout()
plt.savefig("output/hard_voter_30_confusion_matrix_normalised.png", dpi=300)
plt.close()
print(f"{datetime.now()}: Confusion matrix for Hard Voter saved.")
    
"""
plt.figure(figsize=(100, 100))  # Adjust the figure size as necessary

# Mask to only show annotations for non-zero values
mask = np.zeros_like(cm, dtype=bool)
mask[cm == 0] = True

# Create the heatmap
sns.set(font_scale=1.4)  # Adjust font scale as necessary
#ax = sns.heatmap(cm, annot=~mask, fmt="d", cmap="YlGnBu", annot_kws={"size": 16}, mask=mask)
ax = sns.heatmap(cm, annot=~mask, fmt="d", cmap="RdYlGn", xticklabels=test_dataset.features["label"].names, yticklabels=test_dataset.features["label"].names, mask=mask)

# Set labels, title, and ticks
ax.set_xlabel('Predicted labels', fontsize=18)
ax.set_ylabel('True labels', fontsize=18)
ax.set_title('Confusion Matrix', fontsize=20)
ax.tick_params(axis='x', labelrotation=45, labelsize=12)
ax.tick_params(axis='y', labelrotation=0, labelsize=12)

plt.tight_layout()
"""

########################

import pandas as pd
import pandas as pd
import torch
from torch.nn.functional import softmax
import json

# Example placeholders for data
# true_labels = [actual labels for your dataset]
# model_predictions_dict = {'model_name': model_predictions, ...}

# Convert ensemble predictions from logits to class predictions
ensemble_probs = softmax(ensemble_predictions, dim=-1)
ensemble_class_predictions = torch.argmax(ensemble_probs, dim=1).cpu().numpy()

data = {
    'Images' : image_paths,
    'True Label': true_labels.cpu().numpy(),
    'Hard Voter Prediction': ensemble_class_predictions,
}

# Convert each model's predictions from logits to class predictions and add to data
for model_name, preds in model_predictions_dict.items():
    model_probs = softmax(preds, dim=-1)
    model_class_predictions = torch.argmax(model_probs, dim=1).cpu().numpy()
    data[model_name + ' Prediction'] = model_class_predictions

 #Updated function to ensure all NumPy arrays are converted to lists
def to_serializable(val):
    if torch.is_tensor(val):
        return val.cpu().numpy().tolist()  # Convert tensors to lists
    elif isinstance(val, numpy.ndarray):
        return val.tolist()  # Convert numpy arrays to lists
    else:
        return val

# Prepare the data for logging by converting all unserializable objects to lists
logged_data = {key: to_serializable(value) for key, value in data.items()}

# Convert the prepared data to a JSON-formatted string
logged_data_json = json.dumps(logged_data, indent=4)

# Log the data to a JSON file
with open('output/hard_voter_30_logged_data.json', 'w') as f:
    f.write(logged_data_json)

print(f"{datetime.now()}: Results logged.")

# Assuming 'logged_data' is your original dictionary
# And 'id2label' is your mapping from IDs to labels

# Initialize a new dictionary to hold the converted data
converted_data = {}

# Iterate over the items in the original dictionary
for column_name, values in logged_data.items():
    if column_name == 'Images':  # Or whatever the key for the image paths column is
        # Copy the image paths column directly
        converted_data[column_name] = values
    else:
        # Convert each ID to its corresponding label for other columns
        converted_data[column_name] = [id2label[str(id)] for id in values]

# 'converted_data' now has the image paths intact and other columns converted to labels

converted_data["Images"]=sorted(converted_data["Images"])

# Create an empty DataFrame with the desired column names
#columns = ['Image File Name', 'Simple Averager Prediction', 'Model 1 Prediction', 'Model 2 Prediction']
# Add more model columns as necessary
#df = pd.DataFrame(columns=columns)
#df = pd.DataFrame(logged_data)
df = pd.DataFrame(converted_data)

# Save the DataFrame to an Excel file
excel_path='output/hard_voter_30_predictions_visualization.xlsx'
df.to_excel(excel_path, index=False)


print(f"{datetime.now()}: Results saved to {excel_path}")
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load your workbook and select the first sheet
wb = load_workbook("output/hard_voter_30_predictions_visualization.xlsx")
ws = wb.active

# Assuming column B contains the true labels
true_label_col = 'B'
# Assuming column C contains the hard voter predictions, and columns D onwards are individual model predictions
prediction_cols = ['C', 'D', 'E', 'F']  # Add more columns as needed based on your file

# Loop through each row and column to apply formatting
for row in range(2, ws.max_row + 1):  # Skip the header row
    true_label_cell = ws[f'{true_label_col}{row}'].value
    for col in prediction_cols:
        prediction_cell = ws[f'{col}{row}']
        # Apply green fill for matches, red fill for mismatches
        if prediction_cell.value == true_label_cell:
            prediction_cell.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        else:
            prediction_cell.fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")

# Save the workbook to a new file to preserve the original
wb.save("output/hard_voter_30_predictions_visualization_formatted.xlsx")
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Load your workbook and select the first sheet
wb = load_workbook("output/hard_voter_30_predictions_visualization_formatted.xlsx")
ws = wb.active

# Assuming column B contains the true labels
true_label_col = 'B'
# Assuming column C contains the hard voter predictions, and columns D onwards are individual model predictions
prediction_cols = ['C', 'D', 'E', 'F']  # Adjust based on your file

# Initialize a dictionary to hold the correct counts for each prediction column
correct_counts = {col: 0 for col in prediction_cols}

# Loop through each row and column to count correct predictions
for row in range(2, ws.max_row + 1):  # Skip the header row
    true_label_cell_value = ws[f'{true_label_col}{row}'].value
    for col in prediction_cols:
        prediction_cell = ws[f'{col}{row}']
        if prediction_cell.value == true_label_cell_value:
            correct_counts[col] += 1

# Write accuracy information below the last row of data
last_row = ws.max_row + 1
for col in prediction_cols:
    correct = correct_counts[col]
    total = ws.max_row - 1  # Subtract the header row from the total count
    accuracy = (correct / total) * 100

    # Write the counts and accuracy to the next row
    ws[f'{col}{last_row}'] = f"{correct}/{total} Correct"
    ws[f'{col}{last_row + 1}'] = f"{accuracy:.2f}% Accurate"

# Save the workbook
wb.save("output/hard_voter_30_predictions_visualization_with_accuracy.xlsx")

print(f"{datetime.now()}: Script completed.")
