import os
import torch
from datasets import load_dataset
from torch.utils.data import DataLoader
from transformers import AutoModelForImageClassification
from transformers import CvtForImageClassification, DeiTForImageClassification, ViTForImageClassification, ViTHybridForImageClassification, ConvNextV2Model
from transformers import DeiTImageProcessor, ViTImageProcessor, AutoImageProcessor, ViTHybridImageProcessor
import PIL
from PIL import Image
from datetime import datetime
import numpy as np
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import f1_score, confusion_matrix
from sklearn.model_selection import train_test_split
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import json


#####################################################################################################
#### diese Parameter müssen angepasst werden, um das Skript selbst auszuführen ####


#pls select the right split
ds = "/home/student_01/Projects/MA_Cahide/Datasets/Base/20_80_10_10_min4/test"
print(ds)

#Select the Models that you want to use as base for evaluation

model_info = {
    'ViT Base': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/_BaseModels_min4/VisionTransformer/',
        'processor': ViTImageProcessor.from_pretrained('google/vit-base-patch16-224'),
        'custom_class': ViTForImageClassification
    },

    'ViT Large': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/_BaseModels_min4/VisionTransformerLarge/',
        'processor': ViTImageProcessor.from_pretrained('google/vit-large-patch16-384'),
        'custom_class': ViTForImageClassification
    },
    
    'DeiT': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/_BaseModels_min4/DeiT/',
        'processor': DeiTImageProcessor.from_pretrained('facebook/deit-base-distilled-patch16-224'),
        'custom_class': DeiTForImageClassification
    },
    
    'ViTHybrid': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/_BaseModels_min4/ViTHybrid/',
        'processor': ViTHybridImageProcessor.from_pretrained("google/vit-hybrid-base-bit-384"),
        'custom_class': ViTHybridForImageClassification
    },
    
    'CvT': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/_BaseModels_min4/cvt/',
        'processor': AutoImageProcessor.from_pretrained("microsoft/cvt-w24-384-22k") ,
        'custom_class': CvtForImageClassification
    },
    
    'ConvNeXt2': {
        'path': '/home/student_01/Projects/MA_Cahide/Bilderklassifizierung/Modelle/_BaseModels_min4/convnext2/',
        'processor': AutoImageProcessor.from_pretrained("facebook/convnextv2-base-22k-224"),
        'custom_class': AutoModelForImageClassification
    }

}


#####################################################################################################

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

print(f"{datetime.now()}: Using {device} for computation.")

# Load dataset
print(f"{datetime.now()}: ~~~loading dataset~~~")
base_dir = ds
test_dataset = load_dataset('imagefolder', data_dir=base_dir, split="train")
labels = test_dataset.features["label"].names
id2label = {str(i): c for i, c in enumerate(labels)}
label2id = {c: str(i) for i, c in enumerate(labels)}

# Load Models
print(f"{datetime.now()}: ~~~loading models~~~")

output_dir="output/"
os.makedirs(output_dir, exist_ok=True)



# Initialize models and their specific processors
models = {}
for name, info in model_info.items():
    print(f"{datetime.now()}: ~~~loading {name}~~~")
    model_class = info['custom_class'] if 'custom_class' in info else AutoModelForImageClassification
    models[name] = {
        'model': model_class.from_pretrained(info['path']).to(device),
        'processor': info['processor']
    }

image_paths = []
for root, dirs, files in os.walk(base_dir):
    for file in files:
        if file.lower().endswith(('.jpg', '.jpeg')):
            relative_path = os.path.relpath(root, base_dir)
            image_path = os.path.join(relative_path, file)
            image_paths.append(image_path)

def process_image(image_input, processor):
    if isinstance(image_input, str):
        image = Image.open(image_input).convert("RGB")
    elif isinstance(image_input, Image.Image):
        image = image_input.convert("RGB")
    else:
        raise TypeError("The image input must be a file path (str) or a PIL.Image object")
    return processor(images=image, return_tensors="pt")

def get_predictions_for_model(model, dataset, processor):
    model.eval()
    predictions = []
    with torch.no_grad():
        for img_path in dataset['image']:
            processed = process_image(img_path, processor)
            pixel_values = processed['pixel_values'].to(device)
            outputs = model(pixel_values=pixel_values)
            predictions.append(outputs.logits)
    return torch.cat(predictions, dim=0)
"""
# Ensemble prediction logic
model_predictions_dict = {}
for name, info in models.items():
    print(f"{datetime.now()}: Processing {name}")
    predictions = get_predictions_for_model(info['model'], test_dataset, info['processor'])
    model_predictions_dict[name] = predictions

# Logistic Regression Ensemble
print(f"{datetime.now()}: Preparing features for logistic regression.")
ensemble_features = torch.cat(list(model_predictions_dict.values()), dim=1).cpu().numpy()
true_labels = torch.tensor(test_dataset['label']).cpu().numpy()

print(f"{datetime.now()}: Fitting logistic regression model.")
log_reg = LogisticRegression(max_iter=1000)
log_reg.fit(ensemble_features, true_labels)

print(f"{datetime.now()}: Making predictions with logistic regression model.")
ensemble_predictions = log_reg.predict(ensemble_features)

# Calculate accuracy
accuracy = (ensemble_predictions == true_labels).mean()
print(f"{datetime.now()}: Accuracy with Logistic Regression: {accuracy * 100:.2f}%")

# Calculate F1 score
f1 = f1_score(true_labels, ensemble_predictions, average='weighted')
print(f"{datetime.now()}: F1 Score: {f1 * 100:.2f}%")
"""
# Ensemble prediction logic
model_predictions_dict = {}
for name, info in models.items():
    print(f"{datetime.now()}: Processing {name}")
    predictions = get_predictions_for_model(info['model'], test_dataset, info['processor'])
    model_predictions_dict[name] = predictions

# Logistic Regression Ensemble
print(f"{datetime.now()}: Preparing features for logistic regression.")
ensemble_features = torch.cat(list(model_predictions_dict.values()), dim=1).cpu().numpy()
true_labels = torch.tensor(test_dataset['label']).cpu().numpy()

# Log the shape of the features and labels
print(f"Shape of ensemble features: {ensemble_features.shape}")
print(f"Shape of true labels: {true_labels.shape}")

# Split the data into training and validation sets
X_train, X_val, y_train, y_val = train_test_split(ensemble_features, true_labels, test_size=0.2, random_state=42)
print(f"{datetime.now()}: Split data into training and validation sets.")
print(f"Training set size: {len(X_train)}, Validation set size: {len(X_val)}")

# Log a few samples of the training data
print("Sample training features (first 5):")
print(X_train[:5])
print("Sample training labels (first 5):")
print(y_train[:5])

print(f"{datetime.now()}: Fitting logistic regression model.")
log_reg = LogisticRegression(max_iter=1000)
log_reg.fit(X_train, y_train)

print(f"{datetime.now()}: Making predictions with logistic regression model.")
val_predictions = log_reg.predict(X_val)

# Log a few samples of the validation predictions
print("Sample validation predictions (first 5):")
print(val_predictions[:5])
print("Sample validation true labels (first 5):")
print(y_val[:5])

# Calculate accuracy
val_accuracy = (val_predictions == y_val).mean()
print(f"{datetime.now()}: Validation Accuracy with Logistic Regression: {val_accuracy * 100:.2f}%")

# Calculate F1 score
val_f1 = f1_score(y_val, val_predictions, average='weighted')
print(f"{datetime.now()}: Validation F1 Score: {val_f1 * 100:.2f}%")

# Optional: Save intermediate results to a JSON file for further inspection
intermediate_results = {
    "X_train_sample": X_train[:5].tolist(),
    "y_train_sample": y_train[:5].tolist(),
    "val_predictions_sample": val_predictions[:5].tolist(),
    "y_val_sample": y_val[:5].tolist(),
    "val_accuracy": val_accuracy,
    "val_f1": val_f1
}
with open('intermediate_results.json', 'w') as f:
    json.dump(intermediate_results, f, indent=4)
print(f"{datetime.now()}: Intermediate results logged to 'intermediate_results.json'.")

print(f"{datetime.now()}: Making predictions with logistic regression model.")
ensemble_predictions = log_reg.predict(ensemble_features)

# Calculate accuracy
accuracy = (ensemble_predictions == true_labels).mean()
print(f"{datetime.now()}: Accuracy with Logistic Regression: {accuracy * 100:.2f}%")

# Calculate F1 score
f1 = f1_score(true_labels, ensemble_predictions, average='weighted')
print(f"{datetime.now()}: F1 Score: {f1 * 100:.2f}%")

final_decisions_cpu = ensemble_predictions
true_labels_cpu = true_labels
##################################################################################################################################################################
##################################################################################################################################################################

print(f"{datetime.now()}:~~~visualise~~~")

from sklearn.metrics import confusion_matrix
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np
#test_dataset.features["label"].names
# Assuming true_labels and final_decisions are already defined and converted to CPU numpy arrays
cm = confusion_matrix(true_labels_cpu, final_decisions_cpu)


plt.figure(figsize=(30, 30))
mask = np.zeros_like(cm, dtype=bool)
mask[cm == 0] = True
sns.set(font_scale=1.0)
# Uncomment the following line to plot the absolute confusion matrix
# ax = sns.heatmap(cm, annot=mask, fmt="d", cmap="RdYlGn", xticklabels=labels, yticklabels=labels)
ax = sns.heatmap(cm, annot=True, fmt="d", cmap="RdYlGn", mask=mask, cbar=False, xticklabels=labels, yticklabels=labels)
    
ax.set_xlabel('Predicted labels', fontsize=15)
ax.set_ylabel('True labels', fontsize=15)
ax.set_title(f'Confusion Matrix for Hard Voter 10', fontsize=20)
ax.tick_params(axis='x', labelrotation=45, labelsize=10)
ax.tick_params(axis='y', labelrotation=0, labelsize=10)
    
# Save confusion matrix
plt.tight_layout()
plt.savefig("output/log_reg_base_confusion_matrix.png", dpi=300)
plt.close()
print(f"{datetime.now()}: Confusion matrix for Hard Voter saved.")

# Calculate normalized confusion matrix
cm_normalized = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]

# Create a mask to hide the zeroes in the normalized confusion matrix
mask = cm_normalized == 0
    
# Plot normalized confusion matrix
plt.figure(figsize=(30, 30))  # Adjusted figure size for better readability
sns.set(font_scale=1.0)  # Smaller font scale for better readability
ax = sns.heatmap(cm_normalized, annot=~mask, fmt=".0f", cmap="RdYlGn", mask=mask, cbar=True, 
                     xticklabels=labels, yticklabels=labels)
ax.set_xlabel('Predicted labels', fontsize=15)
ax.set_ylabel('True labels', fontsize=15)
ax.set_title(f'Confusion Matrix for Hard Voter', fontsize=18)
ax.tick_params(axis='x', labelrotation=90, labelsize=10)
ax.tick_params(axis='y', labelrotation=0, labelsize=10)
    
# Save normalized confusion matrix
plt.tight_layout()
plt.savefig("output/logreg_base_confusion_matrix_normalised.png", dpi=300)
plt.close()
print(f"{datetime.now()}: Confusion matrix for Hard Voter saved.")
    
"""
plt.figure(figsize=(100, 100))  # Adjust the figure size as necessary

# Mask to only show annotations for non-zero values
mask = np.zeros_like(cm, dtype=bool)
mask[cm == 0] = True

# Create the heatmap
sns.set(font_scale=1.4)  # Adjust font scale as necessary
#ax = sns.heatmap(cm, annot=~mask, fmt="d", cmap="YlGnBu", annot_kws={"size": 16}, mask=mask)
ax = sns.heatmap(cm, annot=~mask, fmt="d", cmap="RdYlGn", xticklabels=test_dataset.features["label"].names, yticklabels=test_dataset.features["label"].names, mask=mask)

# Set labels, title, and ticks
ax.set_xlabel('Predicted labels', fontsize=18)
ax.set_ylabel('True labels', fontsize=18)
ax.set_title('Confusion Matrix', fontsize=20)
ax.tick_params(axis='x', labelrotation=45, labelsize=12)
ax.tick_params(axis='y', labelrotation=0, labelsize=12)

plt.tight_layout()
"""

########################


# Data Logging
data = {
    'Images': image_paths,
    'True Label': true_labels,
    'Logistic Regression Predictions': ensemble_predictions,
}


for model_name, preds in model_predictions_dict.items():
    model_class_predictions = torch.argmax(preds, dim=1).cpu().numpy()
    data[model_name + ' Prediction'] = model_class_predictions

def to_serializable(val):
    if torch.is_tensor(val):
        return val.cpu().numpy().tolist()
    elif isinstance(val, np.ndarray):
        return val.tolist()
    else:
        return val

logged_data = {key: to_serializable(value) for key, value in data.items()}
logged_data_json = json.dumps(logged_data, indent=4)
with open('output/log_reg_base_logged_data.json', 'w') as f:
    f.write(logged_data_json)
print(f"{datetime.now()}: Results logged.")

converted_data = {}
for column_name, values in logged_data.items():
    if column_name == 'Images':
        converted_data[column_name] = values
    else:
        converted_data[column_name] = [id2label[str(id)] for id in values]

converted_data["Images"] = sorted(converted_data["Images"])
df = pd.DataFrame(converted_data)
excel_path = 'output/logreg_base_predictions_visualization.xlsx'
df.to_excel(excel_path, index=False)
print(f"{datetime.now()}: Results saved to {excel_path}")

# Apply formatting to Excel
wb = load_workbook("output/logreg_base_predictions_visualization.xlsx")
ws = wb.active
true_label_col = 'B'
prediction_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I']

for row in range(2, ws.max_row + 1):
    true_label_cell = ws[f'{true_label_col}{row}'].value
    for col in prediction_cols:
        prediction_cell = ws[f'{col}{row}']
        if prediction_cell.value == true_label_cell:
            prediction_cell.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        else:
            prediction_cell.fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
wb.save("output/logreg_base_predictions_visualization_formatted.xlsx")

# Accuracy Calculation and Formatting
wb = load_workbook("output/logreg_base_predictions_visualization_formatted.xlsx")
ws = wb.active
correct_counts = {col: 0 for col in prediction_cols}
for row in range(2, ws.max_row + 1):
    true_label_cell_value = ws[f'{true_label_col}{row}'].value
    for col in prediction_cols:
        prediction_cell = ws[f'{col}{row}']
        if prediction_cell.value == true_label_cell_value:
            correct_counts[col] += 1

last_row = ws.max_row + 1
for col in prediction_cols:
    correct = correct_counts[col]
    total = ws.max_row - 1
    accuracy = (correct / total) * 100
    ws[f'{col}{last_row}'] = f"{correct}/{total} Correct"
    ws[f'{col}{last_row + 1}'] = f"{accuracy:.2f}% Accurate"
wb.save("output/logreg_base_predictions_visualization_with_accuracy.xlsx")
print(f"{datetime.now()}: Script completed.")
